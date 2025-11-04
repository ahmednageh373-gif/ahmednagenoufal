#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=======================================================================
وحدة تحليل الكميات (Quantity Analysis Module)
=======================================================================

الهدف: أتمتة استخلاص وتحليل الكميات من جداول الكميات الأولية (BOQ)
      الناتجة عن AutoCAD/LISP.

المدخلات:
    - جداول الكميات الأولية (BOQ_Initial.xlsx)
    - ملفات DXF (اختياري)

المخرجات:
    - جداول الكميات النهائية (Final_BOQ.xlsx)
    - تقارير حسابات الكميات التفصيلية

الممارسات الهندسية المدمجة:
    - التحقق من كود البناء السعودي (SBC)
    - معايير السلامة الإنشائية
    - منهجية الحساب الواضحة
    
التاريخ: 2025-11-04
المطور: NOUFAL Engineering System
=======================================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging

# إعداد السجلات
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SBCStandards:
    """
    كود البناء السعودي - المعايير الأساسية
    Saudi Building Code - Basic Standards
    """
    
    # SBC 303: Concrete Structures
    CONCRETE_MIN_STRENGTH = {
        'foundations': 25,      # MPa - الأساسات
        'columns': 30,          # MPa - الأعمدة
        'beams': 25,            # MPa - الكمرات
        'slabs': 20,            # MPa - البلاطات
        'retaining_walls': 30   # MPa - الجدران الاستنادية
    }
    
    CONCRETE_MIN_COVER = {
        'cast_against_earth': 75,     # mm - صب مقابل التربة
        'exposed_to_weather': 50,     # mm - معرض للطقس
        'not_exposed_slabs': 20,      # mm - غير معرض - بلاطات
        'not_exposed_beams': 25,      # mm - غير معرض - كمرات
        'not_exposed_columns': 25     # mm - غير معرض - أعمدة
    }
    
    # SBC 303: Steel Reinforcement
    REBAR_MIN_DIAMETER = 8          # mm
    REBAR_MAX_DIAMETER = 40         # mm
    REBAR_MAX_SPACING = {
        'slabs': 300,                # mm
        'beams': 250,                # mm
        'columns': 200               # mm
    }
    
    # Minimum concrete dimensions
    MIN_DIMENSIONS = {
        'slab_thickness': 150,       # mm
        'beam_width': 200,           # mm
        'beam_depth': 300,           # mm
        'column_dimension': 250,     # mm
        'wall_thickness': 150        # mm
    }


class QuantityAnalyzer:
    """
    محلل الكميات الرئيسي
    Main Quantity Analyzer
    """
    
    def __init__(self, project_path: str):
        """
        تهيئة المحلل
        
        Args:
            project_path: المسار الرئيسي للمشروع
        """
        self.project_path = Path(project_path)
        self.input_path = self.project_path / "01_Input_Data" / "03_Cost" / "BOQ_Initial"
        self.output_path = self.project_path / "03_Output_Data" / "01_Quantities"
        self.temp_path = self.project_path / "02_Processing" / "Temp_Data"
        
        # Create output directories if they don't exist
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.temp_path.mkdir(parents=True, exist_ok=True)
        
        self.df_boq = None
        self.validation_results = []
        
        logger.info(f"تم تهيئة محلل الكميات للمشروع: {self.project_path.name}")
    
    def read_boq_data(self, filename: str = "BOQ_Initial.xlsx") -> pd.DataFrame:
        """
        1. قراءة البيانات من جدول الكميات الأولي
        
        Args:
            filename: اسم ملف BOQ
            
        Returns:
            DataFrame: بيانات BOQ
        """
        try:
            file_path = self.input_path / filename
            logger.info(f"قراءة ملف BOQ من: {file_path}")
            
            # Read Excel file
            self.df_boq = pd.read_excel(file_path)
            
            logger.info(f"تم قراءة {len(self.df_boq)} بند من جدول الكميات")
            return self.df_boq
            
        except Exception as e:
            logger.error(f"خطأ في قراءة ملف BOQ: {e}")
            raise
    
    def clean_data(self) -> pd.DataFrame:
        """
        2. تنظيف البيانات
        
        Returns:
            DataFrame: البيانات المنظفة
        """
        if self.df_boq is None:
            raise ValueError("يجب قراءة بيانات BOQ أولاً باستخدام read_boq_data()")
        
        logger.info("بدء تنظيف البيانات...")
        
        # Remove empty rows
        self.df_boq = self.df_boq.dropna(how='all')
        
        # Standardize column names
        column_mapping = {
            'Item No': 'item_number',
            'Item No.': 'item_number',
            'رقم البند': 'item_number',
            'Description': 'description',
            'الوصف': 'description',
            'Unit': 'unit',
            'الوحدة': 'unit',
            'Quantity': 'quantity',
            'الكمية': 'quantity',
            'Length': 'length',
            'الطول': 'length',
            'Width': 'width',
            'العرض': 'width',
            'Height': 'height',
            'الارتفاع': 'height',
            'Thickness': 'thickness',
            'السماكة': 'thickness',
            'Count': 'count',
            'العدد': 'count',
            'Category': 'category',
            'الفئة': 'category',
            'Drawing Ref': 'drawing_ref',
            'مرجع الرسم': 'drawing_ref'
        }
        
        # Rename columns that exist
        for old_name, new_name in column_mapping.items():
            if old_name in self.df_boq.columns:
                self.df_boq = self.df_boq.rename(columns={old_name: new_name})
        
        # Convert numeric columns
        numeric_columns = ['quantity', 'length', 'width', 'height', 'thickness', 'count']
        for col in numeric_columns:
            if col in self.df_boq.columns:
                self.df_boq[col] = pd.to_numeric(self.df_boq[col], errors='coerce')
        
        # Fill NaN values
        self.df_boq = self.df_boq.fillna({
            'length': 1.0,
            'width': 1.0,
            'height': 1.0,
            'thickness': 0.0,
            'count': 1.0,
            'quantity': 0.0
        })
        
        logger.info("تم تنظيف البيانات بنجاح")
        return self.df_boq
    
    def calculate_quantities(self) -> pd.DataFrame:
        """
        3. حساب الكميات النهائية
        
        Returns:
            DataFrame: البيانات مع الكميات المحسوبة
        """
        if self.df_boq is None:
            raise ValueError("يجب تنظيف البيانات أولاً باستخدام clean_data()")
        
        logger.info("بدء حساب الكميات...")
        
        # Calculate final quantity based on unit
        def calculate_final_quantity(row):
            """حساب الكمية النهائية حسب الوحدة"""
            unit = str(row.get('unit', '')).lower()
            
            # Volume (m³)
            if 'm3' in unit or 'm³' in unit or 'متر مكعب' in unit:
                return row['length'] * row['width'] * row['height'] * row['count']
            
            # Area (m²)
            elif 'm2' in unit or 'm²' in unit or 'متر مربع' in unit:
                return row['length'] * row['width'] * row['count']
            
            # Length (m)
            elif 'm' == unit or 'متر' in unit or 'طول' in unit:
                return row['length'] * row['count']
            
            # Count (no, pcs, عدد)
            elif 'no' in unit or 'pcs' in unit or 'عدد' in unit:
                return row['count']
            
            # Default: use existing quantity
            else:
                return row.get('quantity', 0)
        
        # Apply calculation
        self.df_boq['calculated_quantity'] = self.df_boq.apply(calculate_final_quantity, axis=1)
        
        # Add calculation formula column
        def get_calculation_formula(row):
            """توليد صيغة الحساب"""
            unit = str(row.get('unit', '')).lower()
            
            if 'm3' in unit or 'm³' in unit:
                return f"{row['length']:.2f} × {row['width']:.2f} × {row['height']:.2f} × {row['count']:.0f}"
            elif 'm2' in unit or 'm²' in unit:
                return f"{row['length']:.2f} × {row['width']:.2f} × {row['count']:.0f}"
            elif 'm' == unit or 'متر' in unit:
                return f"{row['length']:.2f} × {row['count']:.0f}"
            else:
                return f"{row['count']:.0f}"
        
        self.df_boq['calculation_formula'] = self.df_boq.apply(get_calculation_formula, axis=1)
        
        logger.info(f"تم حساب الكميات لـ {len(self.df_boq)} بند")
        return self.df_boq
    
    def validate_sbc_compliance(self) -> List[Dict]:
        """
        التحقق من الامتثال لكود البناء السعودي
        
        Returns:
            List[Dict]: نتائج التحقق
        """
        if self.df_boq is None:
            raise ValueError("يجب حساب الكميات أولاً")
        
        logger.info("بدء التحقق من الامتثال لكود البناء السعودي...")
        
        self.validation_results = []
        
        for idx, row in self.df_boq.iterrows():
            description = str(row.get('description', '')).lower()
            
            # Check concrete strength
            if 'concrete' in description or 'خرسانة' in description:
                # Extract concrete strength if mentioned
                import re
                strength_match = re.search(r'(\d+)\s*mpa', description, re.IGNORECASE)
                
                if strength_match:
                    strength = int(strength_match.group(1))
                    
                    # Determine element type
                    element_type = None
                    if 'foundation' in description or 'أساس' in description:
                        element_type = 'foundations'
                    elif 'column' in description or 'عمود' in description:
                        element_type = 'columns'
                    elif 'beam' in description or 'كمرة' in description:
                        element_type = 'beams'
                    elif 'slab' in description or 'بلاطة' in description:
                        element_type = 'slabs'
                    
                    if element_type:
                        min_strength = SBCStandards.CONCRETE_MIN_STRENGTH[element_type]
                        
                        if strength < min_strength:
                            self.validation_results.append({
                                'item_number': row.get('item_number'),
                                'description': row.get('description'),
                                'issue': f"قوة الخرسانة ({strength} MPa) أقل من الحد الأدنى المطلوب ({min_strength} MPa) حسب SBC 303",
                                'severity': 'critical',
                                'sbc_code': 'SBC 303'
                            })
            
            # Check slab thickness
            if ('slab' in description or 'بلاطة' in description) and 'thickness' in row:
                thickness = row.get('thickness', 0) * 1000  # convert to mm
                min_thickness = SBCStandards.MIN_DIMENSIONS['slab_thickness']
                
                if thickness > 0 and thickness < min_thickness:
                    self.validation_results.append({
                        'item_number': row.get('item_number'),
                        'description': row.get('description'),
                        'issue': f"سماكة البلاطة ({thickness:.0f} mm) أقل من الحد الأدنى ({min_thickness} mm) حسب SBC 303",
                        'severity': 'critical',
                        'sbc_code': 'SBC 303'
                    })
            
            # Check for drop beams (Architectural preference)
            if ('drop beam' in description or 'كمرة ساقطة' in description):
                self.validation_results.append({
                    'item_number': row.get('item_number'),
                    'description': row.get('description'),
                    'issue': "تم رصد كمرة ساقطة - يُفضل استخدام الكمرات المخفية في المناطق المعمارية الحساسة",
                    'severity': 'warning',
                    'sbc_code': 'Best Practice'
                })
            
            # Check rebar diameter
            if ('rebar' in description or 'حديد' in description or 'تسليح' in description):
                import re
                diameter_match = re.search(r'(\d+)\s*mm', description)
                
                if diameter_match:
                    diameter = int(diameter_match.group(1))
                    
                    if diameter < SBCStandards.REBAR_MIN_DIAMETER:
                        self.validation_results.append({
                            'item_number': row.get('item_number'),
                            'description': row.get('description'),
                            'issue': f"قطر الحديد ({diameter} mm) أقل من الحد الأدنى ({SBCStandards.REBAR_MIN_DIAMETER} mm) حسب SBC 303",
                            'severity': 'critical',
                            'sbc_code': 'SBC 303'
                        })
                    
                    if diameter > SBCStandards.REBAR_MAX_DIAMETER:
                        self.validation_results.append({
                            'item_number': row.get('item_number'),
                            'description': row.get('description'),
                            'issue': f"قطر الحديد ({diameter} mm) أكبر من الحد الأقصى ({SBCStandards.REBAR_MAX_DIAMETER} mm) حسب SBC 303",
                            'severity': 'warning',
                            'sbc_code': 'SBC 303'
                        })
        
        logger.info(f"تم العثور على {len(self.validation_results)} ملاحظة امتثال")
        return self.validation_results
    
    def export_to_excel(self, filename: str = "Final_BOQ.xlsx") -> str:
        """
        4. تصدير الكميات النهائية إلى Excel
        
        Args:
            filename: اسم الملف المخرج
            
        Returns:
            str: مسار الملف المحفوظ
        """
        if self.df_boq is None:
            raise ValueError("لا توجد بيانات للتصدير")
        
        output_file = self.output_path / "Final_BOQ" / filename
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"تصدير البيانات إلى: {output_file}")
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Main BOQ sheet
            self.df_boq.to_excel(writer, sheet_name='BOQ', index=False)
            
            # Validation results sheet
            if self.validation_results:
                df_validation = pd.DataFrame(self.validation_results)
                df_validation.to_excel(writer, sheet_name='SBC Compliance', index=False)
            
            # Summary sheet
            summary_data = {
                'Metric': ['إجمالي البنود', 'البنود الحرجة', 'البنود مع تحذيرات'],
                'Value': [
                    len(self.df_boq),
                    len([r for r in self.validation_results if r['severity'] == 'critical']),
                    len([r for r in self.validation_results if r['severity'] == 'warning'])
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply formatting
        self._format_excel(output_file)
        
        logger.info(f"تم التصدير بنجاح إلى: {output_file}")
        return str(output_file)
    
    def _format_excel(self, filepath: Path):
        """تنسيق ملف Excel"""
        wb = openpyxl.load_workbook(filepath)
        
        # Format BOQ sheet
        if 'BOQ' in wb.sheetnames:
            ws = wb['BOQ']
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format SBC Compliance sheet
        if 'SBC Compliance' in wb.sheetnames:
            ws = wb['SBC Compliance']
            
            # Color code by severity
            for row in ws.iter_rows(min_row=2):
                severity_cell = None
                for cell in row:
                    if cell.column_letter == 'D':  # Severity column
                        severity_cell = cell
                        break
                
                if severity_cell and severity_cell.value == 'critical':
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif severity_cell and severity_cell.value == 'warning':
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        wb.save(filepath)
    
    def run_full_analysis(self, input_filename: str = "BOQ_Initial.xlsx",
                          output_filename: str = None) -> Dict:
        """
        تشغيل التحليل الكامل
        
        Args:
            input_filename: اسم ملف الإدخال
            output_filename: اسم ملف الإخراج (اختياري)
            
        Returns:
            Dict: نتائج التحليل
        """
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Final_BOQ_{timestamp}.xlsx"
        
        logger.info("=" * 70)
        logger.info("بدء التحليل الكامل للكميات")
        logger.info("=" * 70)
        
        # Step 1: Read data
        self.read_boq_data(input_filename)
        
        # Step 2: Clean data
        self.clean_data()
        
        # Step 3: Calculate quantities
        self.calculate_quantities()
        
        # Step 4: Validate SBC compliance
        self.validate_sbc_compliance()
        
        # Step 5: Export to Excel
        output_path = self.export_to_excel(output_filename)
        
        # Prepare results summary
        results = {
            'success': True,
            'output_file': output_path,
            'total_items': len(self.df_boq),
            'total_quantity': self.df_boq['calculated_quantity'].sum(),
            'critical_issues': len([r for r in self.validation_results if r['severity'] == 'critical']),
            'warnings': len([r for r in self.validation_results if r['severity'] == 'warning']),
            'validation_results': self.validation_results
        }
        
        logger.info("=" * 70)
        logger.info("اكتمل التحليل بنجاح!")
        logger.info(f"إجمالي البنود: {results['total_items']}")
        logger.info(f"القضايا الحرجة: {results['critical_issues']}")
        logger.info(f"التحذيرات: {results['warnings']}")
        logger.info(f"ملف الإخراج: {output_path}")
        logger.info("=" * 70)
        
        return results


def main():
    """
    دالة الاختبار الرئيسية
    """
    # Example usage
    project_path = "/home/user/webapp/backend/project_template"
    
    analyzer = QuantityAnalyzer(project_path)
    
    # Create sample BOQ file for testing
    # (في الواقع، سيتم قراءة ملف BOQ من AutoCAD/LISP)
    
    print("=" * 70)
    print("وحدة تحليل الكميات - NOUFAL Engineering System")
    print("=" * 70)
    print("\nالوحدة جاهزة للاستخدام!")
    print("\nمثال على الاستخدام:")
    print("  analyzer = QuantityAnalyzer('/path/to/project')")
    print("  results = analyzer.run_full_analysis('BOQ_Initial.xlsx')")
    print("=" * 70)


if __name__ == "__main__":
    main()
