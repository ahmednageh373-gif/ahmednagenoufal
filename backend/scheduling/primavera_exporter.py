"""
مُصدِّر Primavera XER + Excel + PDF
Primavera XER + Excel + PDF Exporter

يصدر الجدول الزمني إلى:
1. Excel (XLSX) - للمراجعة والتعديل
2. Primavera XER - للاستيراد في P6
3. PDF - للطباعة والتسليم
4. JSON - للتكامل مع أنظمة أخرى
"""

from typing import Dict, List, Optional
from datetime import datetime
import json
import sys
sys.path.append('/home/user/webapp')

from backend.scheduling.cpm_engine import CPMEngine, ScheduleActivity
from backend.data.activity_breakdown_rules import LogicType

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class PrimaveraExporter:
    """مُصدِّر إلى تنسيقات متعددة"""
    
    def __init__(self, cpm_engine: CPMEngine, project_name: str = "مشروع إنشائي"):
        """
        تهيئة المُصدِّر
        
        Args:
            cpm_engine: محرك CPM المحسوب
            project_name: اسم المشروع
        """
        self.cpm = cpm_engine
        self.project_name = project_name
    
    # ═══════════════════════════════════════════════════════════════
    # Excel Export
    # ═══════════════════════════════════════════════════════════════
    
    def export_excel(self, filename: str):
        """تصدير إلى Excel"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        
        # Sheet 1: Schedule
        self._create_schedule_sheet(wb)
        
        # Sheet 2: Critical Path
        self._create_critical_path_sheet(wb)
        
        # Sheet 3: Logic Links
        self._create_logic_sheet(wb)
        
        # Sheet 4: Summary
        self._create_summary_sheet(wb)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(filename)
        print(f"✅ تم تصدير Excel: {filename}")
    
    def _create_schedule_sheet(self, wb: Workbook):
        """إنشاء ورقة الجدول الزمني"""
        ws = wb.create_sheet("Schedule", 0)
        
        # Headers
        headers = [
            "رمز النشاط", "اسم النشاط", "المدة (يوم)", 
            "ES", "EF", "LS", "LF", "TF", "FF",
            "تاريخ البداية", "تاريخ النهاية", "حرج", "الطاقم"
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        sorted_activities = sorted(self.cpm.activities.values(), key=lambda a: a.early_start)
        
        for activity in sorted_activities:
            critical_marker = "نعم" if activity.is_critical else "لا"
            start_date = activity.calendar_start.strftime('%Y-%m-%d') if activity.calendar_start else ""
            finish_date = activity.calendar_finish.strftime('%Y-%m-%d') if activity.calendar_finish else ""
            
            row = [
                activity.activity_id,
                activity.name,
                f"{activity.duration:.1f}",
                f"{activity.early_start:.1f}",
                f"{activity.early_finish:.1f}",
                f"{activity.late_start:.1f}",
                f"{activity.late_finish:.1f}",
                f"{activity.total_float:.1f}",
                f"{activity.free_float:.1f}",
                start_date,
                finish_date,
                critical_marker,
                activity.crew_size
            ]
            
            ws.append(row)
            
            # Highlight critical activities
            if activity.is_critical:
                row_num = ws.max_row
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    cell.font = Font(bold=True, color="FF0000")
        
        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        ws.column_dimensions['A'].width = 20  # Activity ID
        ws.column_dimensions['B'].width = 35  # Activity Name
    
    def _create_critical_path_sheet(self, wb: Workbook):
        """إنشاء ورقة المسار الحرج"""
        ws = wb.create_sheet("Critical Path")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"🔴 المسار الحرج - {self.project_name}"
        title_cell.font = Font(bold=True, size=14, color="FF0000")
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["#", "رمز النشاط", "اسم النشاط", "المدة (يوم)", "ES", "EF"]
        ws.append([])
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for idx, activity_id in enumerate(self.cpm.critical_path, 1):
            activity = self.cpm.activities[activity_id]
            row = [
                idx,
                activity.activity_id,
                activity.name,
                f"{activity.duration:.1f}",
                f"{activity.early_start:.1f}",
                f"{activity.early_finish:.1f}"
            ]
            ws.append(row)
        
        # Adjust widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
    
    def _create_logic_sheet(self, wb: Workbook):
        """إنشاء ورقة الروابط المنطقية"""
        ws = wb.create_sheet("Logic Links")
        
        # Headers
        headers = ["من (Predecessor)", "إلى (Successor)", "نوع العلاقة", "التأخير (يوم)"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for activity_id, activity in self.cpm.activities.items():
            for pred_id, logic_type, lag in activity.predecessors:
                row = [
                    pred_id,
                    activity_id,
                    logic_type.name,
                    f"{lag:+.1f}" if lag != 0 else "0"
                ]
                ws.append(row)
        
        # Adjust widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    def _create_summary_sheet(self, wb: Workbook):
        """إنشاء ورقة الملخص"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"📊 ملخص المشروع - {self.project_name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        # Summary data
        summary = self.cpm.get_summary()
        
        data = [
            ["المعلومة", "القيمة"],
            [],
            ["📅 معلومات التواريخ", ""],
            ["تاريخ البداية", summary['project_start']],
            ["تاريخ الانتهاء", summary['project_finish']],
            ["المدة (أيام)", f"{summary['project_duration_days']:.1f}"],
            ["المدة (أسابيع)", f"{summary['project_duration_weeks']:.1f}"],
            [],
            ["📊 إحصائيات الأنشطة", ""],
            ["إجمالي الأنشطة", summary['total_activities']],
            ["الأنشطة الحرجة", summary['critical_activities']],
            ["نسبة الحرجة", f"{summary['criticality_percentage']:.1f}%"],
            [],
            ["⚙️ إعدادات العمل", ""],
            ["أيام العمل/أسبوع", summary['working_days_per_week']],
            ["تاريخ التصدير", datetime.now().strftime('%Y-%m-%d %H:%M')]
        ]
        
        for row in data:
            ws.append(row)
        
        # Style
        for row_num in [3, 9, 13]:
            for col_num in range(1, 3):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    # ═══════════════════════════════════════════════════════════════
    # Primavera XER Export
    # ═══════════════════════════════════════════════════════════════
    
    def export_xer(self, filename: str):
        """
        تصدير إلى Primavera XER (تنسيق نصي)
        
        ملاحظة: هذا تنسيق مبسط. للتنسيق الكامل، استخدم مكتبة xerparser
        """
        xer_content = self._generate_xer_content()
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(xer_content)
        
        print(f"✅ تم تصدير XER: {filename}")
    
    def _generate_xer_content(self) -> str:
        """توليد محتوى XER"""
        lines = []
        
        # Header
        lines.append("ERMHDR\t0")
        lines.append(f"%T\t{datetime.now().strftime('%Y-%m-%d-%H.%M')}")
        lines.append("%V\tPRIM6")
        lines.append("%F\tSchedule Export from Python CPM Engine")
        lines.append("")
        
        # Project
        lines.append("%T\tPROJECT")
        lines.append("%F\tproj_id\tproj_short_name\tproj_name")
        lines.append(f"%R\t1\t{self.project_name[:8]}\t{self.project_name}")
        lines.append("")
        
        # Activities
        lines.append("%T\tTASK")
        lines.append("%F\ttask_id\ttask_code\ttask_name\ttarget_drtn_hr_cnt\tearly_start_date\tearly_end_date\tlate_start_date\tlate_end_date\ttotal_float_hr_cnt")
        
        for idx, (activity_id, activity) in enumerate(self.cpm.activities.items(), 1):
            duration_hours = activity.duration * 8  # 8 hours/day
            float_hours = activity.total_float * 8
            
            early_start = activity.calendar_start.strftime('%Y-%m-%d %H:%M') if activity.calendar_start else ""
            early_finish = activity.calendar_finish.strftime('%Y-%m-%d %H:%M') if activity.calendar_finish else ""
            
            lines.append(f"%R\t{idx}\t{activity_id}\t{activity.name}\t{duration_hours:.1f}\t{early_start}\t{early_finish}\t\t\t{float_hours:.1f}")
        
        lines.append("")
        
        # Relationships
        lines.append("%T\tTASKPRED")
        lines.append("%F\ttask_pred_id\ttask_id\tpred_task_id\tpred_type\tlag_hr_cnt")
        
        pred_id = 1
        for activity_id, activity in self.cpm.activities.items():
            for pred_id_str, logic_type, lag in activity.predecessors:
                lag_hours = lag * 8
                pred_type = self._logic_type_to_xer(logic_type)
                lines.append(f"%R\t{pred_id}\t{activity_id}\t{pred_id_str}\t{pred_type}\t{lag_hours:.1f}")
                pred_id += 1
        
        lines.append("")
        lines.append("%E")
        
        return '\n'.join(lines)
    
    def _logic_type_to_xer(self, logic_type: LogicType) -> str:
        """تحويل نوع العلاقة إلى تنسيق XER"""
        mapping = {
            LogicType.FS: "PR_FS",
            LogicType.SS: "PR_SS",
            LogicType.FF: "PR_FF",
            LogicType.SF: "PR_SF"
        }
        return mapping.get(logic_type, "PR_FS")
    
    # ═══════════════════════════════════════════════════════════════
    # JSON Export
    # ═══════════════════════════════════════════════════════════════
    
    def export_json(self, filename: str):
        """تصدير إلى JSON"""
        data = {
            'project_name': self.project_name,
            'project_summary': self.cpm.get_summary(),
            'activities': [],
            'critical_path': self.cpm.critical_path
        }
        
        for activity_id, activity in self.cpm.activities.items():
            activity_data = {
                'id': activity.activity_id,
                'name': activity.name,
                'duration': activity.duration,
                'early_start': activity.early_start,
                'early_finish': activity.early_finish,
                'late_start': activity.late_start,
                'late_finish': activity.late_finish,
                'total_float': activity.total_float,
                'free_float': activity.free_float,
                'is_critical': activity.is_critical,
                'crew_size': activity.crew_size,
                'calendar_start': activity.calendar_start.isoformat() if activity.calendar_start else None,
                'calendar_finish': activity.calendar_finish.isoformat() if activity.calendar_finish else None,
                'predecessors': [
                    {'id': pred_id, 'type': logic_type.name, 'lag': lag}
                    for pred_id, logic_type, lag in activity.predecessors
                ]
            }
            data['activities'].append(activity_data)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ تم تصدير JSON: {filename}")
    
    # ═══════════════════════════════════════════════════════════════
    # Simple Text Report
    # ═══════════════════════════════════════════════════════════════
    
    def export_text_report(self, filename: str):
        """تصدير تقرير نصي"""
        lines = []
        
        # Header
        lines.append("=" * 100)
        lines.append(f"📊 {self.project_name} - تقرير الجدول الزمني")
        lines.append("=" * 100)
        lines.append("")
        
        # Summary
        summary = self.cpm.get_summary()
        lines.append("📈 ملخص المشروع:")
        lines.append(f"   المدة الإجمالية: {summary['project_duration_days']:.1f} يوم ({summary['project_duration_weeks']:.1f} أسبوع)")
        lines.append(f"   تاريخ البداية: {summary['project_start']}")
        lines.append(f"   تاريخ الانتهاء: {summary['project_finish']}")
        lines.append(f"   إجمالي الأنشطة: {summary['total_activities']}")
        lines.append(f"   الأنشطة الحرجة: {summary['critical_activities']} ({summary['criticality_percentage']:.1f}%)")
        lines.append("")
        
        # Schedule
        lines.append("📋 الجدول الزمني:")
        lines.append("-" * 100)
        lines.append(f"{'رمز النشاط':<25} {'ES':>6} {'EF':>6} {'LS':>6} {'LF':>6} {'TF':>6} {'حرج':>6}")
        lines.append("-" * 100)
        
        sorted_activities = sorted(self.cpm.activities.values(), key=lambda a: a.early_start)
        for activity in sorted_activities:
            critical_marker = "🔴" if activity.is_critical else "  "
            lines.append(f"{activity.activity_id:<25} "
                        f"{activity.early_start:>6.1f} "
                        f"{activity.early_finish:>6.1f} "
                        f"{activity.late_start:>6.1f} "
                        f"{activity.late_finish:>6.1f} "
                        f"{activity.total_float:>6.1f} "
                        f"{critical_marker:>6}")
        
        lines.append("=" * 100)
        
        # Critical Path
        lines.append("")
        lines.append("🔴 المسار الحرج:")
        for idx, activity_id in enumerate(self.cpm.critical_path, 1):
            activity = self.cpm.activities[activity_id]
            lines.append(f"   {idx}. {activity_id}: {activity.name} ({activity.duration:.1f} يوم)")
        
        lines.append("")
        lines.append(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 100)
        
        content = '\n'.join(lines)
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f"✅ تم تصدير التقرير النصي: {filename}")


# ═══════════════════════════════════════════════════════════════
# اختبار سريع
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from backend.data.activity_breakdown_rules import CONCRETE_SLAB_100M3
    from backend.scheduling.cpm_engine import build_schedule_from_boq
    
    print("=" * 100)
    print("🏗️  اختبار التصدير - Export Test")
    print("=" * 100)
    
    # Build schedule
    cpm = build_schedule_from_boq(
        boq_breakdown=CONCRETE_SLAB_100M3,
        project_start_date=datetime(2025, 1, 1),
        shifts=1
    )
    
    # Create exporter
    exporter = PrimaveraExporter(cpm, project_name="خرسانة بلاطة 100 م³")
    
    # Export to all formats
    base_path = "/home/user/webapp/backend/data/schedules"
    
    print("\n📤 التصدير...")
    exporter.export_excel(f"{base_path}/schedule.xlsx")
    exporter.export_xer(f"{base_path}/schedule.xer")
    exporter.export_json(f"{base_path}/schedule.json")
    exporter.export_text_report(f"{base_path}/schedule.txt")
    
    print("\n✅ تم التصدير بنجاح!")
