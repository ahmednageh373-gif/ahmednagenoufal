#!/usr/bin/env python3
"""
إنشاء خطة المشروع - Create Project Plan
ينشئ ملف Excel يحتوي على خطة تنفيذ شاملة للمشروع
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_project_plan():
    """إنشاء خطة المشروع في ملف Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "خطة المشروع - Project Plan"
    
    # تعريف الأنماط
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12, name="Arial")
    
    phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    phase_font = Font(color="FFFFFF", bold=True, size=11, name="Arial")
    
    task_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    task_font = Font(color="000000", size=10, name="Arial")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العناوين
    headers = [
        "ID",
        "المرحلة / Phase",
        "الوصف / Description",
        "تاريخ البدء / Start Date",
        "تاريخ الانتهاء / End Date",
        "المدة (أيام) / Duration",
        "المسؤول / Owner",
        "الحالة / Status",
        "الملاحظات / Notes"
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # تحديد تاريخ البدء
    start_date = datetime.now()
    
    # البيانات - 5 مراحل رئيسية
    phases = [
        {
            "id": "P1",
            "name_ar": "التحضير والتهيئة",
            "name_en": "Preparation & Setup",
            "duration": 5,
            "owner": "DevOps Team",
            "tasks": [
                {"name_ar": "إعداد قاعدة البيانات PostgreSQL/MongoDB", "name_en": "Setup PostgreSQL/MongoDB", "duration": 2},
                {"name_ar": "تجهيز بيئة التطوير Node.js 18+", "name_en": "Setup Node.js 18+ dev environment", "duration": 1},
                {"name_ar": "إعداد CI/CD و GitHub Actions", "name_en": "Setup CI/CD & GitHub Actions", "duration": 2}
            ]
        },
        {
            "id": "P2",
            "name_ar": "المرحلة الأولى - الأساس الفني",
            "name_en": "Phase 1 - Technical Foundation",
            "duration": 15,
            "owner": "Backend Team",
            "tasks": [
                {"name_ar": "تصميم مخطط قاعدة البيانات (13 جدول)", "name_en": "Design database schema (13 tables)", "duration": 3},
                {"name_ar": "بناء API أساسية لإدارة BOQ والمهام", "name_en": "Build core API for BOQ & tasks", "duration": 5},
                {"name_ar": "تطبيق Authentication (JWT) و RBAC", "name_en": "Implement Authentication & RBAC", "duration": 4},
                {"name_ar": "إنشاء وحدات التكامل الأساسية", "name_en": "Create core integration modules", "duration": 3}
            ]
        },
        {
            "id": "P3",
            "name_ar": "المرحلة الثانية - الواجهة الأمامية",
            "name_en": "Phase 2 - Frontend Development",
            "duration": 12,
            "owner": "Frontend Team",
            "tasks": [
                {"name_ar": "إصلاح أخطاء TypeScript", "name_en": "Fix TypeScript errors", "duration": 2},
                {"name_ar": "بناء مكونات BOQ و Gantt", "name_en": "Build BOQ & Gantt components", "duration": 4},
                {"name_ar": "بناء Dashboard و Progress forms", "name_en": "Build Dashboard & Progress forms", "duration": 4},
                {"name_ar": "ربط الواجهة بـ API", "name_en": "Connect frontend to API", "duration": 2}
            ]
        },
        {
            "id": "P4",
            "name_ar": "المرحلة الثالثة - تكامل الأعمال",
            "name_en": "Phase 3 - Business Integration",
            "duration": 10,
            "owner": "Full-Stack Team",
            "tasks": [
                {"name_ar": "استيراد BOQ من Excel/PDF/CAD", "name_en": "Import BOQ from Excel/PDF/CAD", "duration": 3},
                {"name_ar": "ربط BOQ بالجدول الزمني والتكاليف", "name_en": "Link BOQ to schedule & costs", "duration": 3},
                {"name_ar": "حساب المدة تلقائياً بناءً على الإنتاجية", "name_en": "Auto-calculate duration from productivity", "duration": 2},
                {"name_ar": "تنفيذ حسابات Earned Value", "name_en": "Implement Earned Value calculations", "duration": 2}
            ]
        },
        {
            "id": "P5",
            "name_ar": "المرحلة الرابعة - الاختبار والنشر",
            "name_en": "Phase 4 - Testing & Deployment",
            "duration": 8,
            "owner": "QA & DevOps",
            "tasks": [
                {"name_ar": "كتابة اختبارات الوحدات والتكامل", "name_en": "Write unit & integration tests", "duration": 3},
                {"name_ar": "اختبارات التحميل والأداء", "name_en": "Load & performance testing", "duration": 2},
                {"name_ar": "إنشاء Docker images والنشر", "name_en": "Create Docker images & deploy", "duration": 2},
                {"name_ar": "إعداد النسخ الاحتياطي والمراقبة", "name_en": "Setup backup & monitoring", "duration": 1}
            ]
        }
    ]
    
    # الأنشطة المستمرة
    continuous_activities = [
        {
            "id": "C1",
            "name_ar": "تحسين وتطوير مستمر",
            "name_en": "Continuous Improvement",
            "description": "تفعيل تنبيهات، إضافة AI، تحديث المعايير",
            "owner": "All Teams",
            "status": "مستمر / Ongoing"
        }
    ]
    
    # كتابة البيانات
    current_row = 2
    current_date = start_date
    
    for phase in phases:
        # كتابة المرحلة الرئيسية
        ws.cell(row=current_row, column=1).value = phase["id"]
        ws.cell(row=current_row, column=2).value = f"{phase['name_ar']} / {phase['name_en']}"
        ws.cell(row=current_row, column=3).value = f"مجموعة من {len(phase['tasks'])} مهام"
        ws.cell(row=current_row, column=4).value = current_date.strftime("%Y-%m-%d")
        
        phase_end_date = current_date + timedelta(days=phase["duration"])
        ws.cell(row=current_row, column=5).value = phase_end_date.strftime("%Y-%m-%d")
        ws.cell(row=current_row, column=6).value = phase["duration"]
        ws.cell(row=current_row, column=7).value = phase["owner"]
        ws.cell(row=current_row, column=8).value = "مخطط / Planned"
        
        # تنسيق المرحلة
        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = phase_fill
            cell.font = phase_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        current_row += 1
        
        # كتابة المهام
        task_start_date = current_date
        for i, task in enumerate(phase["tasks"], start=1):
            task_id = f"{phase['id']}.{i}"
            ws.cell(row=current_row, column=1).value = task_id
            ws.cell(row=current_row, column=2).value = ""
            ws.cell(row=current_row, column=3).value = f"{task['name_ar']} / {task['name_en']}"
            ws.cell(row=current_row, column=4).value = task_start_date.strftime("%Y-%m-%d")
            
            task_end_date = task_start_date + timedelta(days=task["duration"])
            ws.cell(row=current_row, column=5).value = task_end_date.strftime("%Y-%m-%d")
            ws.cell(row=current_row, column=6).value = task["duration"]
            ws.cell(row=current_row, column=7).value = phase["owner"]
            ws.cell(row=current_row, column=8).value = "قيد الانتظار / Pending"
            
            # تنسيق المهمة
            for col in range(1, 10):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = task_fill
                cell.font = task_font
                cell.alignment = Alignment(horizontal='left' if col == 3 else 'center', vertical='center')
                cell.border = border
            
            task_start_date = task_end_date
            current_row += 1
        
        current_date = phase_end_date
        current_row += 1  # فراغ بين المراحل
    
    # إضافة الأنشطة المستمرة
    ws.cell(row=current_row, column=1).value = "الأنشطة المستمرة / Continuous Activities"
    for col in range(1, 10):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    current_row += 1
    
    for activity in continuous_activities:
        ws.cell(row=current_row, column=1).value = activity["id"]
        ws.cell(row=current_row, column=2).value = f"{activity['name_ar']} / {activity['name_en']}"
        ws.cell(row=current_row, column=3).value = activity["description"]
        ws.cell(row=current_row, column=4).value = start_date.strftime("%Y-%m-%d")
        ws.cell(row=current_row, column=5).value = "مستمر / Ongoing"
        ws.cell(row=current_row, column=6).value = "N/A"
        ws.cell(row=current_row, column=7).value = activity["owner"]
        ws.cell(row=current_row, column=8).value = activity["status"]
        
        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = task_fill
            cell.font = task_font
            cell.alignment = Alignment(horizontal='left' if col in [2, 3] else 'center', vertical='center')
            cell.border = border
        
        current_row += 1
    
    # تعديل عرض الأعمدة
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    
    # تجميد الصف الأول
    ws.freeze_panes = 'A2'
    
    # حفظ الملف
    filename = f"project_plan_ahmednagenoufal_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print(f"✅ تم إنشاء ملف خطة المشروع: {filename}")
    print(f"📊 عدد المراحل: {len(phases)}")
    print(f"📋 إجمالي المهام: {sum(len(p['tasks']) for p in phases)}")
    print(f"📅 المدة الإجمالية المتوقعة: {sum(p['duration'] for p in phases)} يوم")
    
    return filename

if __name__ == "__main__":
    create_project_plan()
