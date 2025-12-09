#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إصلاح ملف المقايسة - حساب الإجمالي تلقائياً
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys

def fix_boq_file(input_file, output_file):
    """
    قراءة ملف Excel وحساب الإجمالي تلقائياً
    """
    print(f"📂 فتح الملف: {input_file}")
    
    try:
        # فتح الملف
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        print(f"✅ تم فتح الملف بنجاح")
        print(f"📊 عدد الصفوف: {ws.max_row}")
        print(f"📊 عدد الأعمدة: {ws.max_column}")
        
        # البحث عن أعمدة الكمية، سعر الوحدة، والإجمالي
        headers = []
        header_row = 1
        
        # البحث عن صف العناوين
        for row in range(1, min(10, ws.max_row + 1)):
            row_values = [str(cell.value).strip() if cell.value else "" for cell in ws[row]]
            if any(keyword in " ".join(row_values).lower() for keyword in ['كمية', 'quantity', 'سعر', 'price', 'إجمالي', 'total']):
                header_row = row
                headers = row_values
                break
        
        print(f"📋 صف العناوين: {header_row}")
        print(f"📋 العناوين: {headers}")
        
        # تحديد مواقع الأعمدة
        qty_col = None
        price_col = None
        total_col = None
        
        for idx, header in enumerate(headers, start=1):
            header_lower = str(header).lower()
            if 'كمية' in header_lower or 'quantity' in header_lower or 'qty' in header_lower:
                qty_col = idx
            elif 'سعر' in header_lower or 'price' in header_lower or 'rate' in header_lower:
                price_col = idx
            elif 'إجمالي' in header_lower or 'total' in header_lower or 'amount' in header_lower:
                total_col = idx
        
        print(f"📍 عمود الكمية: {qty_col}")
        print(f"📍 عمود السعر: {price_col}")
        print(f"📍 عمود الإجمالي: {total_col}")
        
        if not qty_col or not price_col or not total_col:
            print("❌ لم يتم العثور على الأعمدة المطلوبة!")
            print("🔍 محاولة البحث بطريقة بديلة...")
            
            # محاولة بديلة: البحث عن أرقام في الصفوف
            # افترض أن الأعمدة هي: الوصف، الوحدة، الكمية، السعر، الإجمالي
            for row in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
                row_data = [cell.value for cell in ws[row]]
                print(f"صف {row}: {row_data[:8]}")
                
                # البحث عن أعمدة رقمية
                numeric_cols = []
                for idx, val in enumerate(row_data, start=1):
                    try:
                        if val and float(val) > 0:
                            numeric_cols.append(idx)
                    except:
                        pass
                
                if len(numeric_cols) >= 2:
                    # افترض أن أول عمودين رقميين هما الكمية والسعر
                    qty_col = numeric_cols[0]
                    price_col = numeric_cols[1]
                    # وآخر عمود هو الإجمالي
                    total_col = ws.max_column
                    print(f"✅ تم تحديد الأعمدة: كمية={qty_col}, سعر={price_col}, إجمالي={total_col}")
                    break
        
        if not qty_col or not price_col or not total_col:
            print("❌ فشل في تحديد الأعمدة!")
            return False
        
        # حساب الإجمالي لكل صف
        total_sum = 0
        fixed_count = 0
        
        print(f"\n🔧 بدء إصلاح البنود...")
        
        for row in range(header_row + 1, ws.max_row + 1):
            try:
                # قراءة الكمية والسعر
                qty_cell = ws.cell(row, qty_col)
                price_cell = ws.cell(row, price_col)
                total_cell = ws.cell(row, total_col)
                
                qty = qty_cell.value
                price = price_cell.value
                
                # تحويل إلى أرقام
                try:
                    qty_num = float(qty) if qty else 0
                except:
                    qty_num = 0
                
                try:
                    price_num = float(price) if price else 0
                except:
                    price_num = 0
                
                # حساب الإجمالي
                if qty_num > 0 and price_num > 0:
                    calculated_total = qty_num * price_num
                    
                    # تحديث الخلية
                    total_cell.value = calculated_total
                    total_sum += calculated_total
                    fixed_count += 1
                    
                    # تنسيق الخلية
                    total_cell.number_format = '#,##0.00'
                    
                    if row <= header_row + 5:  # طباعة أول 5 بنود
                        desc = ws.cell(row, 1).value or f"بند {row}"
                        print(f"✅ صف {row}: {desc[:30]} = {qty_num} × {price_num} = {calculated_total:,.2f}")
                
            except Exception as e:
                print(f"⚠️ خطأ في صف {row}: {e}")
                continue
        
        print(f"\n✅ تم إصلاح {fixed_count} بند")
        print(f"💰 الإجمالي الكلي: {total_sum:,.2f} ريال")
        
        # إضافة صف الإجمالي
        summary_row = ws.max_row + 2
        ws.cell(summary_row, 1).value = "الإجمالي الكلي"
        ws.cell(summary_row, 1).font = Font(bold=True, size=14)
        ws.cell(summary_row, total_col).value = total_sum
        ws.cell(summary_row, total_col).number_format = '#,##0.00'
        ws.cell(summary_row, total_col).font = Font(bold=True, size=14)
        ws.cell(summary_row, total_col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # حفظ الملف
        wb.save(output_file)
        print(f"\n✅ تم حفظ الملف المُصلح: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ خطأ: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    input_file = "qassim-contract.xlsx"
    output_file = "qassim-contract-fixed.xlsx"
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    print("="*60)
    print("🔧 برنامج إصلاح ملف المقايسة")
    print("="*60)
    
    success = fix_boq_file(input_file, output_file)
    
    if success:
        print("\n" + "="*60)
        print("✅ تم إصلاح الملف بنجاح!")
        print(f"📁 الملف الأصلي: {input_file}")
        print(f"📁 الملف المُصلح: {output_file}")
        print("="*60)
    else:
        print("\n❌ فشل في إصلاح الملف!")
        sys.exit(1)
