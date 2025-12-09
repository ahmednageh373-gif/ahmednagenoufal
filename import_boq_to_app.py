#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
استيراد ملف المقايسة إلى التطبيق
"""

import openpyxl
import json
import sys
from datetime import datetime

def import_boq_to_json(excel_file, output_json):
    """
    استيراد ملف Excel وتحويله إلى JSON للتطبيق
    """
    print(f"📂 قراءة الملف: {excel_file}")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # البحث عن صف العناوين
        header_row = 6  # من التحليل السابق
        
        # قراءة جميع البنود
        items = []
        categories = {}
        
        for row in range(header_row + 2, ws.max_row + 1):
            try:
                serial = ws.cell(row, 4).value  # الرقم التسلسلي
                category = ws.cell(row, 5).value  # الفئة
                item_code = ws.cell(row, 6).value  # البند
                description = ws.cell(row, 7).value  # وصف البند
                specs = ws.cell(row, 8).value  # المواصفات
                unit = ws.cell(row, 12).value  # وحدة القياس
                quantity = ws.cell(row, 13).value  # الكمية
                unit_price = ws.cell(row, 14).value  # سعر الوحدة
                total = ws.cell(row, 16).value  # الإجمالي
                
                # تحويل إلى أرقام
                try:
                    quantity = float(quantity) if quantity else 0
                except:
                    quantity = 0
                
                try:
                    unit_price = float(unit_price) if unit_price else 0
                except:
                    unit_price = 0
                
                try:
                    total = float(total) if total else (quantity * unit_price)
                except:
                    total = quantity * unit_price
                
                # تخطي البنود الفارغة
                if not description or (quantity == 0 and unit_price == 0):
                    continue
                
                # إنشاء البند
                item = {
                    "id": f"item_{row}",
                    "serialNumber": str(serial) if serial else f"{row}",
                    "category": str(category) if category else "غير محدد",
                    "itemCode": str(item_code) if item_code else "",
                    "description": str(description),
                    "specifications": str(specs) if specs else "",
                    "unit": str(unit) if unit else "م",
                    "quantity": quantity,
                    "unitPrice": unit_price,
                    "total": total,
                    "status": "pending",
                    "progress": 0,
                    "notes": ""
                }
                
                items.append(item)
                
                # إحصائيات الفئات
                cat_name = item["category"]
                if cat_name not in categories:
                    categories[cat_name] = {
                        "name": cat_name,
                        "count": 0,
                        "total": 0
                    }
                categories[cat_name]["count"] += 1
                categories[cat_name]["total"] += total
                
            except Exception as e:
                print(f"⚠️ خطأ في صف {row}: {e}")
                continue
        
        # حساب الإحصائيات
        total_items = len(items)
        total_amount = sum(item["total"] for item in items)
        
        # إنشاء ملف JSON
        data = {
            "projectInfo": {
                "name": "مشروع القصيم",
                "code": "QASSIM-2024",
                "client": "العميل",
                "consultant": "المكتب الاستشاري",
                "contractor": "المقاول",
                "importDate": datetime.now().isoformat(),
                "totalItems": total_items,
                "totalAmount": total_amount
            },
            "summary": {
                "totalItems": total_items,
                "totalQuantity": sum(item["quantity"] for item in items),
                "totalAmount": total_amount,
                "averageUnitPrice": total_amount / sum(item["quantity"] for item in items if item["quantity"] > 0) if any(item["quantity"] > 0 for item in items) else 0
            },
            "categories": list(categories.values()),
            "items": items
        }
        
        # حفظ JSON
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ تم الاستيراد بنجاح!")
        print(f"📊 عدد البنود: {total_items}")
        print(f"📊 عدد الفئات: {len(categories)}")
        print(f"💰 الإجمالي: {total_amount:,.2f} ريال")
        print(f"📁 تم الحفظ في: {output_json}")
        
        # طباعة الفئات
        print(f"\n📋 الفئات:")
        for cat in categories.values():
            print(f"   • {cat['name']}: {cat['count']} بند - {cat['total']:,.2f} ريال")
        
        return True
        
    except Exception as e:
        print(f"❌ خطأ: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_file = "qassim-contract-fixed.xlsx"
    output_json = "qassim-boq-imported.json"
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_json = sys.argv[2]
    
    print("="*60)
    print("📥 برنامج استيراد المقايسة")
    print("="*60)
    
    success = import_boq_to_json(excel_file, output_json)
    
    if success:
        print("\n" + "="*60)
        print("✅ تم الاستيراد بنجاح!")
        print("="*60)
    else:
        print("\n❌ فشل الاستيراد!")
        sys.exit(1)
